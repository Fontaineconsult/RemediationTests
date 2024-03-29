import json
import os

import openpyxl


def create_excel_file(json_file_path, excel_file_path=None):

    with open(json_file_path, 'r') as f:
        data = json.load(f)


    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    wb.remove(ws)
    wb.create_sheet('Documents')
    wb.create_sheet('Document Sites')
    wb.create_sheet('Image Files')
    wb.create_sheet('Video Files')
    wb.create_sheet('Video Sites')
    wb.create_sheet('Audio')
    wb.create_sheet('Audio Sites')
    wb.create_sheet('Unsorted')
    xcel_path = os.path.join(os.path.dirname(json_file_path), data['course_id'] + '.xlsx')
    wb.save(xcel_path)
    build_xcel_file(data, xcel_path)


def find_key_names(d, path=None):
    if path is None:
        path = []
    result = []
    for k, v in d.items():
        new_path = path + [k]
        if isinstance(v, list):
            result.append(new_path)
        elif isinstance(v, dict):
            sub_result = find_key_names(v, new_path)
            result += sub_result
    return result


def add_header_to_sheet(file_path, sheet_name, header_row):
    """
    Adds a header row to a sheet based on the keys in a Python dictionary.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to add the header row to.
        header_row (dict): A dictionary whose keys will be used as the header row.
    """
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Add the header row to the sheet
    for col_idx, header_name in enumerate(sorted(header_row.keys())):
        cell = sheet.cell(row=1, column=col_idx+1)
        cell.value = " ".join(word.capitalize() for word in header_name.replace('_', ' ').split(' '))

    # Save the workbook
    wb.save(file_path)




def add_sheet_to_excel(file_path, sheet_name, rows, headers):

    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    for count, header in enumerate(headers):
        sheet.cell(row=1, column=count).value = header
    for row in rows:
        sheet.append(row)
    wb.save(file_path)



def dicts_to_excel(filename, sheetname, data):
    # Load an existing Excel workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Create a new sheet with the specified sheet name or get the existing one
    if sheetname in wb:
        ws = wb[sheetname]
    else:
        ws = wb.create_sheet(sheetname)

    # Write data rows (values from each dictionary)
    for row_num, item in enumerate(data, 2):  # Start from row 2
        for col_num, key in enumerate(item.keys(), 1):
            ws.cell(row=row_num, column=col_num).value = item.get(key)

    # Save the workbook to a file
    wb.save(filename)



test = r"Z:\ACRS\Cpage\294.json"

def build_xcel_file(json_data, excel_file_path=None):

    key_paths = find_key_names(json_data)
    for path in key_paths:
        # navigate to the list using each path
        sub_dict = json_data
        for key in path[:-1]:
            sub_dict = sub_dict[key]
        my_list = sub_dict[path[-1]]
        print(my_list)
        sheet_name = " ".join(word.capitalize() for word in path[-1].replace('_', ' ').split(' '))

        try:
            add_header_to_sheet(excel_file_path, sheet_name, my_list[0])
            dicts_to_excel(excel_file_path, sheet_name, my_list)
        except IndexError:
            print("IndexError")
            pass


id_list = ['294', '308', '685', '1527', '11998', '12084', '12092', '20349', '20740', '20857', '20935']

for course_id in id_list:
    create_excel_file(fr"Z:\ACRS\Cpage\{course_id}.json")

